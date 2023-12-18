from openpyxl import load_workbook

wb = load_workbook("14-3Dodgers.xlsx")

result = []

ws = wb.worksheets[0]
for row in ws.iter_rows():
    #List comprehension
    result.append([cell.value for cell in row])
    
#累積全壘打sum    
sum = 0
for r in result[1:]:
    sum += int(r[11])
    
print(f"The total number of home runs for Doggers was {sum}.")